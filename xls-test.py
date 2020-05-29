# -*- coding: utf-8 -*-
"""
Created on Fri May 29 20:34:40 2020

@author: Jason Hu
"""

from openpyxl import load_workbook
from openpyxl import Workbook
import os

# 这个地方的地址是你自己电脑上数据的地址
file_path = r"F:\教程\openpyxl\website.xlsx"
new_file_path = r"F:\教程\openpyxl\new_website.xlsx"
output_path = r"F:\教程\openpyxl\output"

# 加载工作簿
wb = load_workbook(file_path )
# 获取活动中的表格
ws = wb.active
print("open sheet:", ws.title)
print("row:", ws.max_row, "column:", ws.max_column)

# 从第二行开始
for row in range(2, ws.max_row + 1):
    name = ws.cell(row, 1).value # 获取名字
    site = ws.cell(row, 2).value # 获取网址
    print("name:", name, "site:", site)
    with open(output_path + "\\" + name + ".txt", "w") as fp:
        fp.write(site)
print("output file ok.")

new_wb = Workbook() # 创建一个新的工作簿
# 创建一个新的工作表格
new_ws = new_wb.create_sheet("result")
print("new sheet:", new_ws.title)
# 设置第一行的标题
new_ws["A1"] = "名字"
new_ws["B1"] = "网址"
row = 2 # 从第2行开始写入
# 列出目录下面的所有文件
files = os.listdir(output_path)
for file in files:
    print(file)
    with open(output_path + "\\" + file) as fp:     
        print(file.split("."))
        new_ws.cell(row, 1, file.split(".")[0]) # 设置名字
        new_ws.cell(row, 2, fp.read())  # 设置网址
        row += 1

new_wb.save(new_file_path)  # 保存工作簿
new_wb.close() # 关闭工作簿