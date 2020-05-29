# -*- coding: utf-8 -*-
"""
Created on Fri May 29 15:20:10 2020

@author: Jason Hu
"""

#import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

save_path = r"F:\教程\openpyxl" + r"\test.xlsx"

print("> create a new work book:")
# 创建一个工作簿
wb = Workbook()

print("> select sheet:")
# 创建工作表
wb.create_sheet("sheet1")
wb.create_sheet("sheet2")

# 名字方式选择
ws = wb["sheet1"]
print(ws.title)

# 索引方式选择
ws = wb.worksheets[0]
print(ws.title)
ws = wb.worksheets[1]
print(ws.title)

# 选择激活的工作表
ws = wb.active
print(ws.title)

print("> write data:")

# 往层里面写入数据
for row in range(1, 5):
    for column in range(1, 5):
        ws.cell(row, column).value = str(row) + ',' + str(column)
        ws.cell(row+5, column, str(row) + ',' + str(column))

# 通过字符设置数据
ws["A1"] = 123
ws["A2"] = "test"
ws["B1"] = 1 + 2
ws["A1"].value = 123
ws["A2"].value = "test"
ws["B1"].value = 1 + 2

# 通过字符获取数据
print(ws["A1"].value)
print(ws["A2"].value)
print(ws["B1"].value)

# 通过切片获取范围
cell_range = ws["A1":"C3"]
print(cell_range)
for row in cell_range:
    for cell in row:
        print(cell.value)
        
print("> save work book:")
wb.save(save_path)  # 保存文件
wb.close()          # 关闭工作簿

print("> load a existed work book:")
# 加载工作簿
wb = load_workbook(save_path)

print("> print work book info:")

# 选择激活的工作表
ws = wb.active
print(ws.title)

print("row:", ws.max_row, "column:", ws.max_column)

for row in range(1, ws.max_row + 1):    
    for column in range(1, ws.max_column + 1):
        print(ws.cell(row, column).value, end=' ')
    print('')

# 查看所有工作表
print(wb.sheetnames)
